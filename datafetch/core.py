"""FUNCTIONS FROM INDUSTRIAL DATA SCIENCE WORKFLOW (IDSW) PACKAGE
Pipelines for Google Cloud Platform (GCP), Google Colab and Google Drive
Pipelines for Amazon Simple Storage Service (S3)
Pipelines for reading tables and non-structured sheets on Excel
Pipelines for loading Pandas dataframe and exporting as CSV and Excel
Extract data from Plant Information Management (PIMS) systems
AspenTech IP21
Connect to SQLite Database

Marco Cesar Prado Soares, Data Scientist Specialist @ Bayer Crop Science LATAM
marcosoares.feq@gmail.com
marco.soares@bayer.com"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl

from dataclasses import dataclass
from colabutils import (InvalidInputsError, ControlVars)


@dataclass
class Connectors:
  """
    Store the connectors as Global variables to use, so that the classes do not have to be re-instantiated always.
    Variables are always started with value None or as booleans.

    for a variable var, syntax 
    if var:
        # is equivalent to: "run only when variable var is not None." or "run only if var is True".
    """
  
  # While persistent = True, the system will try to use the same connector already created.
  # User must change this variable state to create new connectors.
  persistent = True
  
  # This class will also store the connectors, once they are created.


class MountGoogleDrive:
    """Class for Mounting Google Drive"""
    def __init__ (self):
        pass

    def mount_drive(self):

        from google.colab import drive
        # Google Colab library must be imported only in case it is
        # going to be used, for avoiding AWS compatibility issues.
        
        print("Associate the Python environment to your Google Drive account, and authorize the access in the opened window.")
        
        drive.mount('/content/drive')
        
        print("Now your Python environment is connected to your Google Drive: the root directory of your environment is now the root of your Google Drive.")
        print("In Google Colab, navigate to the folder icon (\'Files\') of the left navigation menu to find a specific folder or file in your Google Drive.")
        print("Click on the folder or file name and select the elipsis (...) icon on the right of the name to reveal the option \'Copy path\', which will give you the path to use as input for loading objects and files on your Python environment.")
        print("Caution: save your files into different directories of the Google Drive. If files are all saved in a same folder or directory, like the root path, they may not be accessible from your Python environment.")
        print("If you still cannot see the file after moving it to a different folder, reload the environment.")
    

    def upload_to_colab(self):

        from google.colab import files
        # google.colab library must be imported only in case 
        # it is going to be used, for avoiding 
        # AWS compatibility issues.
        print("Click on the button for file selection and select the files from your machine that will be uploaded in the Colab environment.")
        print("Warning: the files will be removed from Colab memory after the Kernel dies or after the notebook is closed.")
        # this functionality requires the previous declaration:
        ## from google.colab import files
            
        colab_files_dict = files.upload()
            
        # The files are stored into a dictionary called colab_files_dict where the keys
        # are the names of the files and the values are the files themselves.
        ## e.g. if you upload a single file named "dictionary.pkl", the dictionary will be
        ## colab_files_dict = {'dictionary.pkl': file}, where file is actually a big string
        ## representing the contents of the file. The length of this value is the size of the
        ## uploaded file, in bytes.
        ## To access the file is like accessing a value from a dictionary: 
        ## d = {'key1': 'val1'}, d['key1'] == 'val1'
        ## we simply declare the key inside brackets and quotes, the same way we would do for
        ## accessing the column of a dataframe.
        ## In this example, colab_files_dict['dictionary.pkl'] access the content of the 
        ## .pkl file, and len(colab_files_dict['dictionary.pkl']) is the size of the .pkl
        ## file in bytes.
        ## To check the dictionary keys, apply the method .keys() to the dictionary (with empty
        ## parentheses): colab_files_dict.keys()
            
        for key in colab_files_dict.keys():
            #loop through each element of the list of keys of the dictionary
            # (list colab_files_dict.keys()). Each element is named 'key'
            print(f"User uploaded file {key} with length {len(colab_files_dict[key])} bytes.")
            # The key is the name of the file, and the length of the value
            ## correspondent to the key is the file's size in bytes.
            ## Notice that the content of the uploaded object must be passed 
            ## as argument for a proper function to be interpreted. 
            ## For instance, the content of a xlsx file should be passed as
            ## argument for Pandas .read_excel function; the pkl file must be passed as
            ## argument for pickle.
            ## e.g., if you uploaded 'table.xlsx' and stored it into colab_files_dict you should
            ## declare df = pd.read_excel(colab_files_dict['table.xlsx']) to obtain a dataframe
            ## df from the uploaded table. Notice that is the value, not the key, that is the
            ## argument.
                
            print("The uploaded files are stored into a dictionary object named as colab_files_dict.")
            print("Each key from this dictionary is the name of an uploaded file. The value correspondent to that key is the file itself.")
            print("The structure of a general Python dictionary is dict = {\'key1\': value1}. To access value1, declare file = dict[\'key1\'], as if you were accessing a column from a dataframe.")
            print("Then, if you uploaded a file named \'table.xlsx\', you can access this file as:")
            print("uploaded_file = colab_files_dict[\'table.xlsx\']")
            print("Notice, though, that the object uploaded_file is the whole file content, not a Python object already converted. To convert to a Python object, pass this element as argument for a proper function or method.")
            print("In this example, to convert the object uploaded_file to a dataframe, Pandas pd.read_excel function could be used. In the following line, a df dataframe object is obtained from the uploaded file:")
            print("df = pd.read_excel(uploaded_file)")
            print("Also, the uploaded file itself will be available in the Colaboratory Notebook\'s workspace.")
            
            self.colab_files_dict =  colab_files_dict

            return self
    

    def download_from_colab(self, file_to_download_from_colab = None):

        from google.colab import files

            
        if (file_to_download_from_colab is None):
                
            #No object was declared
            print("Please, inform a file to download from the notebook\'s workspace. It should be declared in quotes and with the extension: e.g. \'table.csv\'.")
            
        else:
                
            print("The file will be downloaded to your computer.")

            files.download(file_to_download_from_colab)

            print(f"File {file_to_download_from_colab} successfully downloaded from Colab environment.")
        
        return self


class SQLServerConnection:
    """
    Class for extracting data from a SQL Server instance.
    def __init__ (self, server, 
                  database,
                  username = '', 
                  password = '',
                  system = 'windows'):
    
    : param: system = 'windows', 'macos' or 'linux'

        If the user passes the argument, use them. Otherwise, use the standard values.
        Set the class objects' attributes.
        Suppose the object is named assistant. We can access the attribute as:
        assistant.assistant_startup, for instance.
        So, we can save the variables as objects' attributes.

    INSTALL ODBC DRIVER IF USING MACOS OR LINUX (Windows already has it):
    - MacOS Installation: https://learn.microsoft.com/en-us/sql/connect/odbc/linux-mac/install-microsoft-odbc-driver-sql-server-macos?view=sql-server-ver16
    - Linux Installation: https://learn.microsoft.com/en-us/sql/connect/odbc/linux-mac/installing-the-microsoft-odbc-driver-for-sql-server?view=sql-server-ver16&tabs=alpine18-install%2Calpine17-install%2Cdebian8-install%2Credhat7-13-install%2Crhel7-offline
    """

    # Initialize instance attributes.
    # define the Class constructor, i.e., how are its objects:

    def __init__ (self, server, 
                  database,
                  username = '', 
                  password = '',
                  system = 'windows'):

        
        import pyodbc
        # Some other example server values are
        # server = 'localhost\sqlexpress' # for a named instance
        # server = 'myserver,port' # to specify an alternate port

        if ((username is None)|(username == '')):
            # Ask the user to provide the credentials:
            username = input(f"Enter your username for accessing the SQL database {database} from server {server} here (in the right).")
            print("\n") # line break
        
        if ((password is None)|(password == '')):
            from getpass import getpass
            password = getpass(f"Enter your password (Secret key) for accessing the SQL database {database} from server {server} here (in the right).")
        
        
        if (system == 'windows'):
            cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
        
        else:
            cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password + 'Encrypt=no;TrustServerCertificate=yes')
            # https://stackoverflow.com/questions/71587239/operationalerror-when-trying-to-connect-to-sql-server-database-using-pyodbc/71588236#71588236
        
        cursor = cnxn.cursor()
        
        self.cnxn = cnxn
        self.cursor = cursor
        self.query_counter = 0
        

    def get_db_schema (self, show_schema = True, export_csv = False, saving_directory_path = "db_schema.csv"):
        """
        : param: show_schema (bool): if True, the schema of the tables on the SQL Server will be shown.
        """
            
        query = "SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE='BASE TABLE'"
            
        schema_df = pd.read_sql(query, self.cnxn)
        
        if ControlVars.show_results: # dominant context
            if (show_schema):
                print("Database schema:")
                print(f"There are {len(schema_df)} tables registered in the database.\n")
                try:
                    from IPython.display import display
                    display(schema_df)
                    
                except:
                    print(schema_df)
                
        self.schema_df = schema_df

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = "db_schema.csv"
            
            schema_df.to_csv(saving_directory_path)
            
        return self
        
        
    def run_sql_query (self, query, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """

        # SQL SERVER INNER JOIN:
        """
        https://www.sqlservertutorial.net/sql-server-basics/sql-server-inner-join/

        FROM table1 d
        INNER JOIN table2 t
        ON d.key1 = t.key2
        """
        
        # SQL SERVER UNION STATEMENT (vertical concatenation/append):
        """
        https://www.w3schools.com/sql/sql_union.asp

        FROM table1 d
        INNER JOIN table2 t
        ON d.key1 = t.key2
        UNION
        SELECT ValueTime as timestamp, TagName AS tag
        FROM LatestIP21TagDataNumeric
        WHERE TagName = 'TAGABC' OR TagName = 'TAGABD' OR TagName = 'TAGABE' 
        ORDER BY timestamp ASC;
        -- Notice that ORDER BY appears at the end of the query, after all joins and unions
        """

        # SQL SERVER TOP (EQUIVALENT TO LIMIT):
        """
        https://www.w3schools.com/sqL/sql_top.asp
        
        SELECT TOP 100 *
        FROM table;
        """
        
        # SQL SERVER WHERE STATEMENT (FILTER):
        """
        https://learn.microsoft.com/en-us/sql/t-sql/queries/where-transact-sql?view=sql-server-ver16

        WHERE column1 = 'val1' OR column1 = 'val2'
        """
        
        # SQL SERVER CASE STATEMENT (IF ELSE):
        """
        https://www.w3schools.com/sql/sql_case.asp

        CASE
            WHEN t.TagName = 'CODEXXX' THEN 'Variable X'
        ELSE ''
        END AS variable,
        """

        query_counter = self.query_counter
            
        df = pd.read_sql(query, self.cnxn)

        if ControlVars.show_results: # dominant context
            if (show_table):   
                print("Returned table:\n")
                try:
                    from IPython.display import display
                    display(df)
                    
                except:
                    print(df)

        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_query{query_counter}"] = df

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df.to_csv(saving_directory_path)
        
        # Update counter:
        self.query_counter = query_counter + 1

        return df
        
        
    def get_full_table (self, table, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: table (str): string containing the name of the table that will be queried.
        """
        
        query_counter = self.query_counter

        query = "SELECT * FROM " + str(table)
            
        df_table = pd.read_sql(query, self.cnxn)
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned table:\n")
                try:
                    from IPython.display import display
                    display(df_table)
                    
                except:
                    print(df_table)
        
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_table{query_counter}"] = df_table

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df_table.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1    
            
        return df_table
    
    
    def query_specific_tag_ip21sqlserver (self, tag, variable_name = None, show_table = True, export_csv = False, saving_directory_path = ""):
        """ : param: tag (str): string with tag as registered in IP21. e.g. tag = 'ABC00AA101-01'.
        
            : param: variable_name (str): string containing a more readable name for the tag, that will be also shown.
            e.g. variable_name = 'Temperature in C'
        """
        
        # https://www.sqlservertutorial.net/sql-server-basics/sql-server-inner-join/
        # https://www.w3schools.com/sqL/sql_top.asp
        # https://learn.microsoft.com/en-us/sql/t-sql/queries/where-transact-sql?view=sql-server-ver16
        # https://www.w3schools.com/sql/sql_case.asp
        # https://www.w3schools.com/sql/sql_union.asp
        
        if (variable_name is None):
            #Repeat the tag
            variable_name = tag

        query = f"""SELECT d.ValueTime AS timestamp, t.TagName AS tag,
                    CASE
                        WHEN t.TagName = '{tag}' THEN '{variable_name}'
                        ELSE ''
                    END AS variable,
                    d.Value AS value
                    FROM IP21DataNumeric d
                    INNER JOIN IP21PublishConfig t
                    ON d.TagConfigID = t.ID
                    WHERE t.TagName = '{tag}'
                    UNION
                    SELECT ValueTime as timestamp, TagName AS tag, 
                    CASE
                        WHEN TagName = '{tag}' THEN '{variable_name}'
                        ELSE ''
                    END AS variable,
                    Value AS value
                    FROM LatestIP21TagDataNumeric
                    WHERE TagName = '{tag}'
                    ORDER BY timestamp ASC;
                """
        
        tag_df = self.run_sql_query(query = query, show_table = show_table, export_csv = export_csv, saving_directory_path = saving_directory_path)
        
            
        return tag_df


class SQLiteConnection:
    """Class for connecting and manipulating a SQLite Database file"""

    def __init__(self, file_path, pre_created_engine = None):

        self.file_path = file_path
        self.engine = pre_created_engine
    

    def create_engine(self):
        
        # Make imports and create the engine for the database
        # Configure the SQLite engine
        from sqlalchemy import create_engine
        
        # SQLAlchemy engines documentation
        # https://docs.sqlalchemy.org/en/20/core/engines.html
        # SQLite connects to file-based databases, using the Python built-in module sqlite3 by default.
        # As SQLite connects to local files, the URL format is slightly different. 
        # The “file” portion of the URL is the filename of the database. For a relative file path, this requires 
        # three slashes:
        # sqlite://<nohostname>/<path>
        # where <path> is relative:

        try:
                    
            if (self.file_path[:2] == './'):
                # Add a slash, since sqlite engine requires 3 slashes
                self.file_path = self.file_path[2:]
                
            if (self.file_path[0] == '/'):
                # Add a slash, since sqlite engine requires 3 slashes
                self.file_path = self.file_path[1:]
                        
            self.file_path = """sqlite:///""" + self.file_path
            # file_path = "sqlite:///my_db.db"
                    
            self.engine = create_engine(self.file_path)
            #And for an absolute file path, the three slashes are followed by the absolute path:
                
            """
            # Unix/Mac - 4 initial slashes in total
            engine = create_engine("sqlite:////absolute/path/to/foo.db")
                
            # Windows
            engine = create_engine("sqlite:///C:\\path\\to\\foo.db")
                
            # Windows alternative using raw string
            engine = create_engine(r"sqlite:///C:\path\to\foo.db")
            To use a SQLite :memory: database, specify an empty URL:
                
            engine = create_engine("sqlite://")
            More notes on connecting to SQLite at SQLite.
            """
        
        except:
            raise InvalidInputsError ("Error trying to create SQLite Engine Database. Check if no more than one slash was added to file path.\n")
        
        return self
    

    def fetch_table(self, table_name):
        
        # If there is no engine, create one:
        if (self.engine is None):
            self = self.create_engine()
        
        try:
            # Access the table from the database
            df = pd.read_sql(table_name, self.engine)
            
            if ControlVars.show_results: 
                print(f"Successfully retrieved table {table_name} from the database.")
                print("Check the 10 first rows of the dataframe:\n")
                
                try:
                    # only works in Jupyter Notebook:
                    from IPython.display import display
                    display(df.head(10))
                        
                except: # regular mode
                    print(df.head(10))
            
            return df, self.engine
        
        except:
            raise InvalidInputsError ("Error trying to fetch SQLite Database. If an pre-created engine was provided, check if it is correct and working.\n")
        

    def update_or_create_table(self, table_name):
    
        # If there is no engine, create one:
        if (self.engine is None):
            self = self.create_engine()
            
        try:
            # Set index = False not to add extra indices in the database:
            df.to_sql(table_name, con = engine, if_exists = 'replace', index = False)
                
            if ControlVars.show_results: 
                print(f"Successfully updated table {table_name} on the SQLite database.")
                print("Check the 10 first rows from this table:\n")
                    
                try:
                    # only works in Jupyter Notebook:
                    from IPython.display import display
                    display(df.head(10))
                            
                except: # regular mode
                    print(df.head(10))

            return df, self.engine
        
        except:
            raise InvalidInputsError ("Error trying to update SQLite Database. If an pre-created engine was provided, check if it is correct and working.\n")
            

class GCPBigQueryConnection:
    """
    Class for accessing Google Cloud Platform (GCP) BigQuery data.

    : param: project (str): project name on BigQuery
    : param: dataset (str): dataset that user wants to connect

    : param: already_authenticated (bool): True if the manual connection to GCP was already performed
        and so the user is authorized.
        
        This connection can be done by running the following command directly on a cell or on the console
        from Python environment. Attention: follow the command line authentication instruction below.

            !gcloud auth application-default login
        
    
        COMMAND LINE AUTHENTICATION INSTRUCTION
            
            - BEFORE INSTATIATING THE CLASS, FOLLOW THESE GUIDELINES!
            
            1. Copy and run the following line in a notebook cell. Do not add quotes.
            The "!" sign must be added to indicate the use of a command line software:

                !gcloud auth application-default login
            
            2. Also, you can run the SQL query on GCP console and, when the query results appear, click on EXPLORE DATA - Explore with Python notebook.
            It will launch a Google Colab Python notebook that you may run for data exploring, manipulation and exporting.

            2.1. To export data, you expand the hamburger icon on the left side of the Google Colab, click on Files, and then select an exported CSV or other files. Finally, click on the ellipsis (3 dots) and select Download to obtain it.
            

        Install Google Cloud Software Development Kit (SDK) before running
                
        General Instructions for Installation: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#installing_the_latest_version
        Instructions for Windows: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#windows
        Instructions for Mac OS: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#mac
        Instructions for Linux: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#linux
        
        From: https://stackoverflow.com/questions/39419754/downloading-and-importing-google-cloud-python
        First, make sure you have installed gcloud on your system then run the commands like this:

        First: gcloud components update in your terminal.
        then: pip install google-cloud
        And for the import error:
        Adding "--ignore-installed" to pip command may work.
        This might be a bug in pip - see this page for more details: https://github.com/pypa/pip/issues/2751

        This pipeline may be blocked also due to security configurations, and may fail on some Virtual Private Networks (VPNs).
        - Try to run this pipeline outside of the VPN in case it fails

    """
    
    # Initialize instance attributes.
    # define the Class constructor, i.e., how are its objects:

    def __init__ (self, project = '', dataset = '', already_authenticated = True):
       

        from google.cloud import bigquery
        from google.cloud import bigquery_storage
        from google.oauth2 import service_account
        import google.auth
        
        if ((project is None)|(project == '')):
            # Ask the user to provide the credentials:
            # This is the name that appears on the right-hand menu from GCP console Big Query page, 
            # (https://console.cloud.google.com/bigquery) called Explorer
            # that contains a group of datasets. E.g.: location360-datasets; bcs-csw-core, etc
            # The individual datasets are revealed after expanding the project name by clicking on the arrow.
            print("\n")
            self.project = input(f"Enter the name of the project registered on Google Cloud Platform (GCP).\n")
            
        if ((dataset is None)|(dataset == '')):
            # Ask the user to provide the credentials:
            # E.g.: core
            print("\n")
            self.dataset = input(f"Enter the name of the dataset from project {self.project} registered on GCP, containing the tables that will be queried.\n")

        self.query_counter = 0
        self.already_authenticated = already_authenticated
        

    def authenticate (self, authentication_method = 'manual',
                  vault_secret_path = '', app_role = '', app_secret = ''):
        """
        : param: authentication_method (str): 'manual' or 'vault' authentication
            : param: authetication_method = 'manual' for GCP standard manual authentication on browser. System will try
                to access the authorization window, in case it was not done yet.
            : param: authetication_method = 'vault' for Vault automatic authentication, dependent on corporate
                cibersecurity and data asset.

        : param: vault_secret_path (str): path to access the secret
        : params: vault_secret_path = '', app_role = '', app_secret = '' are the parameters for vault authorization
            
        """
        
        if self.already_authenticated:
            self.bqclient = bigquery.Client(project = self.project)
            try:
                self.bqstorageclient = bigquery_storage.BigQueryReadClient()
            except:
                self.bqstorageclient = None # create an empty attribute
        
        else:

            if (authentication_method == 'manual'):
                self = self.manual_authentication()
            
            elif (authentication_method == 'vault'):
                self = self.vault_authentication(vault_secret_path = vault_secret_path, app_role = app_role, app_secret = app_secret)

        return self


    def manual_authentication (self):

        from google.cloud import bigquery
        from google.cloud import bigquery_storage
        from google.oauth2 import service_account
        import google.auth
        
        try:
            # Setting authorization from cloud. Might seem no effect if CLI is configured
            self.credentials, self.project = google.auth.default(
                scopes = ["https://www.googleapis.com/auth/cloud-platform"]) 
            
            self.bqclient = bigquery.Client(credentials = self.credentials, project = self.project)
            self.bqstorageclient = bigquery_storage.BigQueryReadClient(credentials = self.credentials)
        
        except:
            # Some other example server values are
            # server = 'localhost\sqlexpress' # for a named instance
            # server = 'myserver,port' # to specify an alternate port
            
            # subprocess module allows running shell commands. Each portion from a Bash script is declared as an element 
            # from a list of strings. Outputs are captured as a list of strings as well.        
            # Example:
            """
            with Popen(["ls"], stdout=PIPE) as proc:
            out = proc.readlines()
            print(out)
            
            output: ['some_file.txt','some_other_file.txt']
            # Notice that a Python variable of string type may be included to the list for generating dynamic execution of commands.
            """

            """
            Example 2 command: python -m pip show pandas
            from subprocess import Popen, PIPE, TimeoutExpired
            proc = Popen(["python", "-m", "pip", "show", "pandas"], stdout = PIPE, stderr = PIPE)
            try:
                output, error = proc.communicate(timeout = 15)
                print(output)
            except:
                        # General exception
                output, error = proc.communicate()
                print(f"Process with output: {output}, error: {error}.\n")

            output: b'Name: pandas\r\nVersion: 2.0.3\r\nSummary: Powerful data structures for data analysis, time series, and statistics\r\nHome-page: \r\nAuthor: \r\nAuthor-email: The Pandas Development Team <pandas-dev@python.org>\r\nLicense: BSD 3-Clause License\r\n        \r\n        Copyright (c) 2008-2011, AQR Capital Management, LLC, Lambda Foundry, Inc. and PyData Development Team\r\n        All rights reserved.\r\n        \r\n        Copyright (c) 2011-2023, Open source contributors.\r\n        \r\n        Redistribution and use in source and binary forms, with or without\r\n        modification, are permitted provided that the following conditions are met:\r\n        \r\n        * Redistributions of source code must retain the above copyright notice, this\r\n          list of conditions and the following disclaimer.\r\n        \r\n        * Redistributions in binary form must reproduce the above copyright notice,\r\n          this list of conditions and the following disclaimer in the documentation\r\n          and/or other materials provided with the distribution.\r\n        \r\n        * Neither the name of the copyright holder nor the names of its\r\n          contributors may be used to endorse or promote products derived from\r\n          this software without specific prior written permission.\r\n        \r\n        THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS"\r\n        AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE\r\n        IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE\r\n        DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE\r\n        FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL\r\n        DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR\r\n        SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER\r\n        CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY,\r\n        OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE\r\n        OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.\r\n        \r\nLocation: c:\\users\\gnklm\\appdata\\local\\anaconda3\\lib\\site-packages\r\nRequires: numpy, python-dateutil, pytz, tzdata\r\nRequired-by: cmdstanpy, datashader, holoviews, hvplot, prophet, seaborn, shap, statsmodels, xarray\r\n'
            """

            from subprocess import Popen, PIPE, TimeoutExpired

            # Start a long running process using subprocess.Popen()
            # Command to run:
            """!gcloud auth application-default login"""

            try:
                proc = Popen(["gcloud", "auth", "application-default", "login"], stdout = PIPE, stderr = PIPE)

                """
                You will use the subprocess.communicate() method to wait for the command to finish running for up to 15 seconds. 
                The process will then timeout and it will return an Exception: i.e. error detected during execution, which will be caught 
                and the process will be cleaned up by proc.kill(). 
                """

                # Use subprocess.communicate() to create a timeout 
                try:
                    output, error = proc.communicate(timeout = 15)
                    # Simply remove timeout argument if process is not supposed to finish after a given time.

                except TimeoutExpired:

                    # Cleanup the process if it takes longer than the timeout
                    proc.kill()

                    # Read standard out and standard error streams and print
                    output, error = proc.communicate()
                    print(f"Process timed out with output: {output}, error: {error}.")

                except:
                    # General exception
                    proc.kill()
                    output, error = proc.communicate()
                    print(f"Process of Google Cloud Mount with output: {output}, error: {error}.\n")

                    warning = """
                    Install Google Cloud Software Development Kit (SDK) before running this function.

                    General Instructions for Installation: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#installing_the_latest_version
                    Instructions for Windows: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#windows
                    Instructions for Mac OS: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#mac
                    Instructions for Linux: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#linux
                    """

                    print(warning)

                """try:
                    from IPython.display import display 
                    # from IPython.display import display, display_html
                    out = list(proc._fileobj2output.keys())
                    out1, out2 = out[1], out[0]
                    display(proc._fileobj2output[out1][0])
                    print("\n")
                    # display(proc._fileobj2output[out2][0])
                    # display_html(proc._fileobj2output[out2][0])

                except:
                    pass"""
            
            except:
                pass
        
            try:
                self.bqclient = bigquery.Client(project = self.project)
            except:

                error_msg = """
                Impossible to Connect with the input parameters or protocols. 
                Try manual connection running

                    !gcloud auth application-default login
                
                on a cell or on the Python IDE's terminal.
                """
                raise InvalidInputsError(error_msg)


        return self


    def get_vault_secret (self, vault_secret_path: str, app_role: str, app_secret: str):
               
        import base64
        import json
        import hvac
        import os

        from google.oauth2 import service_account
        from google.cloud import bigquery, bigquery_storage
        
        """
        Get the vault secret in dictionary
        More detail about hvac: https://hvac.readthedocs.io/en/stable/overview.html
        
        """
        vault_url = 'https://vault.agro.services'
        vault_client = hvac.Client(url = vault_url)
        
        vault_client.auth.approle.login(app_role, app_secret)

        vault_path = vault_secret_path
        vault_secret = vault_client.read(vault_path)
        
        return vault_secret['data']['data']

 
    def get_vault_credentials (self, vault_secret_path: str, app_role: str, app_secret: str):
        
        import base64
        import json
        import hvac
        import os

        from google.oauth2 import service_account
        from google.cloud import bigquery, bigquery_storage
        
        """
        Bigquery project credential with service account
        Return credential of a service account using app role and app secret
        
        """
        vault_secret = self.get_vault_secret(vault_secret_path, app_role, app_secret)

        if 'data' in vault_secret and type(vault_secret['data']) == str:
            service_account_creds = json.loads(base64.b64decode(vault_secret['data']))
        
        else:
            # in case credentials are saved directly as json object in vault (not encoded) you can get it directly
            service_account_creds = json.loads(base64.b64decode(vault_secret))
        
        bq_credentials = service_account.Credentials.from_service_account_info(service_account_creds)

        return bq_credentials


    def vault_authentication (self, vault_secret_path = '', app_role = '', app_secret = ''):     
        
        import base64
        import json
        import hvac
        import os

        from google.oauth2 import service_account
        from google.cloud import bigquery, bigquery_storage

        
        if ((vault_secret_path is None)|(vault_secret_path == '')):
            vault_secret_path = input(f"Enter Vault Secret Path for accessing CSW.")
            print("\n") # line break
        
        if ((app_role is None)|(app_role == '')):
            app_role = input(f"Enter App Role for getting the Vault client.")
            print("\n") # line break

        if ((app_secret is None)|(app_secret == '')):
            from getpass import getpass
            vault_secret_path = getpass(f"Enter App Secret for getting the Vault client.")
            print("\n") # line break

        self.credentials = self.get_vault_credentials(vault_secret_path = vault_secret_path, app_role = app_role, app_secret = app_secret)

        self.bqclient = bigquery.Client(credentials = self.credentials, project = self.project)
        self.bqstorageclient = bigquery_storage.BigQueryReadClient(credentials = self.credentials)

        return self


    def table_exists (self, table_id: str) -> bool:
        """
        Checks if a table with the specified ID exists in BigQuery.

        Parameters:
        - client (bigquery.Client): The BigQuery client.
        - table_id (str): The ID of the table to check.

        Returns:
        bool: True if the table exists, False otherwise.
        """
        client = self.bqclient

        try:
            client.get_table(table_id)
            return True
        
        except:
            return False


    def run_sql_query (self, query, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """

        query_counter = self.query_counter

        client = self.bqclient
        job = client.query(query)
        df = job.to_dataframe()

        if ControlVars.show_results: # dominant context
            if (show_table):   
                print("Returned table:\n")
                try:
                    from IPython.display import display
                    display(df)
                    
                except:
                    print(df)

        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_query{query_counter}"] = df

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df.to_csv(saving_directory_path)
        
        # Update counter:
        self.query_counter = query_counter + 1

        return df
        
        
    def get_full_table (self, table, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: table (str): name of the table to be retrieved. Full table name is `{self.project}.{self.dataset}.{str(table)}`
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """

        table_name = f"""`{self.project}.{self.dataset}.{str(table)}`"""
        
        query_counter = self.query_counter

        query = "SELECT * FROM " + table_name
            
        client = self.bqclient
        job = client.query(query)
        df_table = job.to_dataframe()
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned table:\n")
                try:
                    from IPython.display import display
                    display(df_table)
                    
                except:
                    print(df_table)
        
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_table{query_counter}"] = df_table

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df_table.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1    
            
        return df_table
    

    def write_data_on_bigquery_table (self, table, df):
        """
        : param: table (str): string with table name
        : param: df (pd.DataFrame): Pandas dataframe to be written on BigQuery table
        """

        client = self.bqclient
        table_ref = client.dataset(self.dataset).table(str(table))
        table = client.get_table(table_ref)

        errors = client.insert_rows_from_dataframe(table, df)

        if errors != [[]]: 
            raise RuntimeError('Error in writing to BigQuery: {}'.format(errors))

        else:
            
            print(f"Dataframe written on Big Query {table} table from dataset {self.project}.{self.dataset}.\n")
            
            return self


    def delete_specific_values_from_column_on_table (self, table, column, values_to_delete, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: column (str): is the column name on a given BigQuery table (a string).
        : param: values_to_delete is a single value (numeric or string) or an iterable containing a set
          of values to be deleted.
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """
        
        # column is the column name on a given BigQuery table (a string).
        # values_to_delete is a single value (numeric or string) or an iterable containing a set
        # of values to be deleted.


        if (type(values_to_delete) == str):
            # put inside list:
            values_to_delete = [values_to_delete]
        
        else:
            try:
                values_to_delete = list(values_to_delete)
            
            except:
                # It is not an iterable (it is a value):
                values_to_delete = [values_to_delete]

        # pick 1st element and remove it from list
        val0 = values_to_delete.pop(0)

        if (type(val0) == str):
            
            delete_query = f"""
                                DELETE 
                                    FROM `{self.project}.{self.dataset}.{str(table)}`
                                    WHERE {column} = '{val0}' """
        
        else:
            delete_query = f"""
                                DELETE 
                                    FROM `{self.project}.{self.dataset}.{str(table)}`
                                    WHERE {column} = {val0} """

        # Now, loop through the remaining list (if there is a remaining one) to update query.
        # If list is empty, the loop does not run
        for val in values_to_delete:
            if (type(val) == str):
                delete_query = delete_query + f"""OR {column} = '{val}' """
                
            else:
                delete_query = delete_query + f"""OR {column} = {val} """

        client = self.bqclient
        query_counter = self.query_counter

        table_ref = client.dataset(self.dataset).table(str(table))
        table = client.get_table(table_ref)

        table.streaming_buffer

        if table.streaming_buffer is None:
            job = client.query(delete_query)
            job.result()
        else:
            raise RuntimeError("Table contains data in a streaming buffer, which cannot be updated or deleted. Please perform this action when the streaming buffer is empty (may take up to 90 minutes from last data insertion).")

        df_table = job.to_dataframe()
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned table with deleted data:\n")
                try:
                    from IPython.display import display
                    display(df_table)
                    
                except:
                    print(df_table)
        
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_table{query_counter}"] = df_table

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df_table.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1    
            
        return df_table    


    def update_specific_value_from_column_on_table (self, table, column, old_value, updated_value, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: column (str): is the column name on a given BigQuery table (a string).
        : param: old_value: value that must be replaced
        : param: updated_value: new value to be added.
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """
            
        if ((type(old_value) == str)|(type(updated_value) == str)):
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = '{updated_value}'
                        WHERE `{column}` = '{old_value}'
                        """
            
        else:
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = {updated_value}
                        WHERE `{column}` = {old_value}
                        """
            
        client = self.bqclient
        query_counter = self.query_counter

        table_ref = client.dataset(self.dataset).table(str(table))
        table = client.get_table(table_ref)

        table.streaming_buffer

        if table.streaming_buffer is None:
            job = client.query(update_query)
            job.result()
        else:
            raise RuntimeError("Table contains data in a streaming buffer, which cannot be updated or deleted. Please perform this action when the streaming buffer is empty (may take up to 90 minutes from last data insertion).")

        df_table = job.to_dataframe()
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned table with updated data:\n")
                try:
                    from IPython.display import display
                    display(df_table)
                    
                except:
                    print(df_table)
        
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_table{query_counter}"] = df_table

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df_table.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1    
            
        return df_table


    def update_entire_column_from_table (self, table, column, updated_value, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: column (str): is the column name on a given BigQuery table (a string).
        : param: updated_value: new value to be added.
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """
            
        if (type(updated_value) == str):
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = '{updated_value}'
                        WHERE TRUE
                        """
            
        else:
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = {updated_value}
                        WHERE TRUE
                        """
            
        client = self.bqclient
        query_counter = self.query_counter

        table_ref = client.dataset(self.dataset).table(str(table))
        table = client.get_table(table_ref)

        table.streaming_buffer

        if table.streaming_buffer is None:
            job = client.query(update_query)
            job.result()
        else:
            raise RuntimeError("Table contains data in a streaming buffer, which cannot be updated or deleted. Please perform this action when the streaming buffer is empty (may take up to 90 minutes from last data insertion).")

        df_table = job.to_dataframe()
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned table with updated data:\n")
                try:
                    from IPython.display import display
                    display(df_table)
                    
                except:
                    print(df_table)
        
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_table{query_counter}"] = df_table

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df_table.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1    
            
        return df_table    


    def update_value_when_finding_str_or_substring_on_another_column (self, table, column, updated_value, string_column, str_or_substring_to_search, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: column (str): is the column name on a given BigQuery table (a string).
        : param: updated_value: new value to be added.
        : param: string_column (str): column containing a string or substring that will be searched.
        : param: str_or_substring_to_search (str): (in quotes): string or substring that will be searched on 
            column 'string_column'. When it is find, the value on 'column' will be updated.
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """
        
        if (type(updated_value) == str):
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = '{updated_value}'
                        WHERE CONTAINS_SUBSTR({string_column}, "{str_or_substring_to_search}")
                        """
            
        else:
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = {updated_value}
                        WHERE CONTAINS_SUBSTR({string_column}, "{str_or_substring_to_search}")
                        """
            
        client = self.bqclient
        query_counter = self.query_counter

        table_ref = client.dataset(self.dataset).table(str(table))
        table = client.get_table(table_ref)

        table.streaming_buffer

        if table.streaming_buffer is None:
            job = client.query(update_query)
            job.result()
        else:
            raise RuntimeError("Table contains data in a streaming buffer, which cannot be updated or deleted. Please perform this action when the streaming buffer is empty (may take up to 90 minutes from last data insertion).")

        df_table = job.to_dataframe()
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned table with updated data:\n")
                try:
                    from IPython.display import display
                    display(df_table)
                    
                except:
                    print(df_table)
            
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_table{query_counter}"] = df_table

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df_table.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1    
            
        return df_table


    def update_value_when_finding_numeric_value_on_another_column (self, table, column, updated_value, comparative_column, value_to_search, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: column (str): is the column name on a given BigQuery table (a string).
        : param: updated_value: new value to be added.
        : param: comparative_column (str): column containing a numeric value that will be searched.
        : param: value_to_search: numeric value that will be searched 
            on column 'comparative_colum'. When it is find, the value on 'column' will be updated.
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """
        
        if (type(updated_value) == str):
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = '{updated_value}'
                        WHERE `{comparative_column}` = {value_to_search}
                        """
            
        else:
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = {updated_value}
                        WHERE `{comparative_column}` = {value_to_search}
                        """
            
        client = self.bqclient
        query_counter = self.query_counter

        table_ref = client.dataset(self.dataset).table(str(table))
        table = client.get_table(table_ref)

        table.streaming_buffer

        if table.streaming_buffer is None:
            job = client.query(update_query)
            job.result()
        else:
            raise RuntimeError("Table contains data in a streaming buffer, which cannot be updated or deleted. Please perform this action when the streaming buffer is empty (may take up to 90 minutes from last data insertion).")

        df_table = job.to_dataframe()
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned table with updated data:\n")
                try:
                    from IPython.display import display
                    display(df_table)
                    
                except:
                    print(df_table)
        
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_table{query_counter}"] = df_table

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df_table.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1    
            
        return df_table
    

    def create_new_view (self, view_id, query, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        Creates a view in Google BigQuery if a view with the same name does not already exist.
        
        Important! This function uses a connection from google cloud sdk already configured

        Parameters:
        : param: view_id (str): The ID of the view to be created. If no ID is provided, a table is created
        : param: query (str): The SQL query defining the view.

        Returns:
        None
        """

        from google.cloud import bigquery

        view_id, query = str(view_id), str(query)
        # self.project = project
        bqclient, bqstorageclient = self.bqclient, self.bqstorageclient
        
        # Check if the view already exists
        if self.table_exists(bqclient, view_id):
            print(f"A view with the ID '{view_id}' already exists. Not creating a new view.")
            return self

        
        df = (
            bqclient.query(query)
            .result()
            .to_dataframe()
        )

        # This step creates the table for specified view
        view = bigquery.Table(view_id)
        view.view_query = query

        # Make an API request to create the view.
        view = bqclient.create_table(view)

        query_counter = self.query_counter
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned view:\n")
                try:
                    from IPython.display import display
                    display(df)
                    
                except:
                    print(df)
        
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_view{query_counter}"] = df

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1 

        return df


class IngestExcelTables:
    """
    Class for picking Excel files with non-structured data saved into several tables.
    For this class to work, the tables do not need to be input in a structured format, but they must be formatted
    as tables. With that, they can be detected and converted to Pandas dataframes. Alternatively, if no table is
    detected, the whole sheet is loaded

    : param: file_path (str): full path where the Excel file is locally stored. The file extension must be provided 
    (xlsx, etc).

    """
    
    def __init__ (self, file_path):
        
        self.file_path = file_path
        # Load workbook Openpyxl documentation:
        # https://openpyxl.readthedocs.io/en/stable/worksheet_tables.html?highlight=load_workbook#table-as-a-print-area
        self.wb = openpyxl.load_workbook(file_path)
        self.worksheets = self.wb.worksheets # list of worksheets
        self.loaded_dfs = [] # list of loaded dataframes
    
    
    def pre_cleansing(self, df):
        """Pre-cleansing of the dataframe. The methods are highly prone to result in duplicate rows and 
        completely blank rows or columns"""
        
        # Remove completely blank columns:
        df = df.dropna(axis = 1, how = 'all')
        # Remove completely blank rows:
        df = df.dropna(axis = 0, how = 'all')
        # Drop duplicates:
        df = df.drop_duplicates()
        # Reset index:
        df = df.reset_index(drop = True)
        
        return df
    
    
    def get_table_parameters(self, tab_range):
        
        """Use the table ranges identified by Openpyxl to obtain the
        parameters skiprows and usecols from pd.read_excel function.
        
        skiprows (int): indicates row much rows to skip.
        usecols (str): indicates the columns to use, in format "A:B"
        
        The tab_range obtained with openpyxl, in turns, is a string as "A10:BC78", which
        may contain an arbitrary number of letters, followed by an arbitrary number of digits.
        """
        
        # Split string as "A10:BC78" in ":". It will create a list with two strings ["A10", "BC78"]
        ranges = tab_range.split(":")
        
        # Iterate through characters in the first string from list ranges
        for i in range(len(ranges[0])):
            try:
                # If it was possible to break the string in two and convert the second part to
                # integer, thus the process may be finished:
                first_col, first_row = ranges[0][:i], int(ranges[0][i:])
                break
            
            except:
                # the error is raised if it was not possible to convert to int (e.g. 'A1') is not
                # conversible. So pass, and try again.
                pass
        
        # Repeat the process for the second string
        for i in range(len(ranges[1])):
            try:
                last_col, last_row = ranges[1][:i], int(ranges[1][i:])
                break
            
            except:
                pass
        
        # Concatenate strings to obtain usecols parameter:
        usecols = first_col + ":" + last_col
        
        # Parameter skipcols will be one unit less tha first_row
        # For instance, "A10" starts on the tenth row, so 9 rows must be skipped:
        skiprows = first_row - 1
        
        return usecols, skiprows

        
    def read_table(self, sheet_name, tab_range, has_header = True):
        
        # Get table parameters for pd.read_excel function:
        usecols, skiprows = self.get_table_parameters(tab_range)
        
        # Read table as pandas dataframe:
        
        if (has_header == True):         
            table = pd.read_excel(self.file_path, sheet_name = sheet_name, skiprows = skiprows, usecols = usecols, na_values = None, verbose = False, parse_dates = True)
        
        else:
            table = pd.read_excel(self.file_path, sheet_name = sheet_name, header = None, skiprows = skiprows, usecols = usecols, na_values = None, verbose = False, parse_dates = True)
        
        # Do the pre-cleansing:
        table = self.pre_cleansing(table)
            
        return table
    
    
    def read_full_sheet(self, sheet_name, has_header = True):
        """Read the entire sheet, instead of an individual table"""
        
        if (has_header == True):         
            table = pd.read_excel(self.file_path, sheet_name = sheet_name, na_values = None, verbose = False, parse_dates = True)

        else:
            table = pd.read_excel(self.file_path, sheet_name = sheet_name, header = None, na_values = None, verbose = False, parse_dates = True)
                    
        # Pre-cleansing:
        table = self.pre_cleansing(table)
        
        return table

    
    def load_dfs(self, has_header = True):
        
        loaded_dfs = self.loaded_dfs
        
        for ws in self.worksheets:
            sheet_name = ws.title
            
            # ws.tables is a dictionary containing the tables in a given sheet. If the dictionary is empty,
            # (len = 0), load the whole sheet as a dataframe:
            if (len(ws.tables) == 0):
                table = self.read_full_sheet(sheet_name, has_header)
                # Store on the list:
                loaded_dfs.append({'sheet': sheet_name, 'table': sheet_name, 'df': table})
                
            
            else:
                # Loop through each table:
                for table_name, tab_values in zip(ws.tables, ws.tables.values()):
                    tab_range = tab_values.ref

                    # Read the table:
                    table = self.read_table(sheet_name, tab_range, has_header)
                    # Store on the list:
                    loaded_dfs.append({'sheet': sheet_name, 'table': table_name, 'df': table})
        
        
        # Save list as attribute:
        self.loaded_dfs = loaded_dfs
                
        return self
    
    
    def export_processed_excel_file(self):
        
        file_path = "processed_excel.xlsx"
        
        try:
            # The replacement of a Sheet will only occur in the append ('a') mode.
            # 'a' is a mode available for the cases where an Excel file is already present.
            # Let's check if there is an Excel file previously created, so that we will not
            # delete it:
            with pd.ExcelWriter(file_path, date_format = "YYYY-MM-DD",
                                datetime_format = "YYYY-MM-DD HH:MM:SS",
                                mode = 'a', if_sheet_exists = 'replace') as writer:
                
                for loaded_df in self.loaded_dfs:
                    df, table = loaded_df['df'], loaded_df['table']
                    
                    if (table != loaded_df['sheet']):
                        table = loaded_df['sheet'] + "_" + table
                        
                    df.to_excel(writer, sheet_name = table, na_rep='', 
                                    header = True, index = False, 
                                    startrow = 0, startcol = 0, merge_cells = False, 
                                    inf_rep = 'inf')
        
        except:
            # The context manager created by class ExcelWriter with 'a' mode returns an error when
            # there is no Excel file available. Since we do not have the risk of overwriting the file,
            # we can open the writer in write ('w') mode to create a new spreadsheet:
            with pd.ExcelWriter(file_path, date_format = "YYYY-MM-DD",
                                datetime_format = "YYYY-MM-DD HH:MM:SS", mode = 'w') as writer:   
                
                for loaded_df in self.loaded_dfs:
                    df, table = loaded_df['df'], loaded_df['table']
                    
                    if (table != loaded_df['sheet']):
                        table = loaded_df['sheet'] + "_" + table
                        
                    df.to_excel(writer, sheet_name = table, index = False, 
                                startrow = 0, startcol = 0, merge_cells = False, 
                                inf_rep = 'inf')
        
        
        if ControlVars.show_results: 
            print(f"Dataframes exported as Excel file to notebook\'s workspace as \'{file_path}\'.")
            print("Warning: if there was a sheet with the same name as the exported ones, it was replaced by the exported dataframe.")
        
        
    def ingestion_pipeline(self, has_header = True, export_excel = True):
        
        self = self.load_dfs(has_header)
        
        if export_excel:
            self.export_processed_excel_file()
        
        return self


class SharePointDownloader:
    """Pipeline for accessing SharePoint files
    This class provides a reusable tool for connecting to SharePoint and manipulating dataframes sourced from it.

    The behavior of the class revolves around six main points:
    1. The initialization method gathers essential information, except for two pieces of information required during the download method call:
        - CLIENT_ID, CLIENT_SECRET, and TENANT_ID: These are from the Azure APP created, necessary for the application to run successfully. They should be set as environment variables with secret encapsulation.
        - SITE_NAME: This variable is set for each call, representing the SharePoint site to be accessed. Without it, the URL requesting (get() method) cannot proceed.
        - COMPANY_TENANT_ID: This variable sets the company tenant ID, similar to the initial part of the SharePoint URL (e.g., your_company.sharepoint.com).
     
        You can identify it in your url address as: 
                                
                                your_company.sharepoint.com
    
    : param: company_tenant_id (str): you may pass it directly to the constructor as a string. example:
        sharepointdownloader = SharePointDownloader(company_tenant_id = 'your_company'). Alternatively, keep
        company_tenant_id = None to try to collect it from the environment.
    
    The class provides five main functions:
    1. get_token: Retrieves the header information required for each get() method.
    2. get_response_id: Searches for the ID for every folder/file level, crucial for constructing the URL to be accessed.
    3. get_drive_id: Retrieves the drive ID information from the first level of folders in the SharePoint library.
    4. find_file: Searches for the desired file within each folder on the pipeline. It avoids looping the same search after finding each level of subfolder.
    5. download_file: Initiates the entire pipeline. This method requires setting two variables:
        - target_file_name: The name of the file, along with its type (e.g., 'file.xlsx' for Excel, 'file.csv' for CSV).
        - folder_match: The folder name to be matched, to find the root folder ID.
    """

    def __init__(self, ccompany_tenant_id = None, client_id = None, client_secret = None, 
                tenant_id = None, site_name = None):
        """
        Initializes the SharepointDownloader object.

        params: 
            - company_tenant_id (str): you may pass it directly to the constructor as a string. 
            - client_id (str): you may pass it directly to the constructor as a string. 
            - client_secret (str): you may pass it directly to the constructor as a string. 
            - tenant_id (str): you may pass it directly to the constructor as a string. 
            - site_name (str): you may pass it directly to the constructor as a string. 
            
        example:
            sharepointdownloader = SharePointDownloader(
                                                        company_tenant_id = 'your_company', 
                                                        client_id = 'you client ID from Azure',
                                                        client_secret = 'you client secret from Azure', 
                                                        tenant_id = 'you tenant ID from Azure',
                                                        site_name = 'The sharepoint site name as shown at the sharepoint's URL',
                                                        ). 
            Alternatively, keep company_tenant_id = None to try to collect it from the environment (or any/all other variables).
        """

        import os
        from dotenv import load_dotenv
        from msal import ConfidentialClientApplication
        import requests

        load_dotenv()

         # Environment variables loading
         # If None, use the second option (os.environ.get method)
        self.COMPANY_TENANT_ID = company_tenant_id or os.environ.get('COMPANY_TENANT_ID')   
        self.CLIENT_ID = client_id or os.environ.get('CLIENT_ID')
        self.CLIENT_SECRET = client_secret or os.environ.get('CLIENT_SECRET')
        self.TENANT_ID = tenant_id or os.environ.get('TENANT_ID')
        self.SITE_NAME = site_name or os.environ.get('SITE_NAME')
            
        # Authentication Config
        self.authority = 'https://login.microsoftonline.com/' + self.TENANT_ID
        self.scope = ['https://graph.microsoft.com/.default']

        # APP Initialization
        self.app = ConfidentialClientApplication(
            self.CLIENT_ID, 
            authority = self.authority,
            client_credential = self.CLIENT_SECRET
        )
        
        # This variable is retrieved after the Env variableS "site name" and "company tenant (like your_company.sharepoint.com)" is set, dynamically searching the sharepoint site ID.
        self.SHAREPOINT_SITE_ID = requests.get(f'https://graph.microsoft.com/v1.0/sites/{self.COMPANY_TENANT_ID}:/sites/{self.SITE_NAME}', headers = self.get_token()).json()['id']
       
        # This url is used at each request, as it is the main request url
        self.url = f'https://graph.microsoft.com/v1.0/sites/{self.SHAREPOINT_SITE_ID}/drives/'
    

    def get_token(self):
        """
        This function aims to create a 'headers' object to be used at each get request on the Graph API.
        The return is a json/dictionary object to be read by the headers argument, as it can be seen in the functions below.
        """
        from dotenv import load_dotenv
        load_dotenv()

        try:
            result = self.app.acquire_token_for_client(scopes = self.scope)
            access_token = result['access_token']
            headers = {'Authorization': 'Bearer ' + access_token}
            return headers
        
        except Exception as e:
            print(f'Error: {e}')
        
    
    def get_response_id(self, result_json, folder_match):
        """
            This function is responsible to read all API response and return the folder/file IDs.
            It converts the response from get() method into a JSON format and gets the intended ID by matching the subsequent name on arg.
            
            The arguments of this function are:
            - result_json: inputs the obtject as type requests.models.Response, which is converted to json and read by the function
            - folder_match: string of folder/file to be found by the function. 
            
            If the name is not match as required, the return is None.
        """

        import json
        from dotenv import load_dotenv
        load_dotenv()

        try:
            # Converting the response into JSON format
            result_json = json.loads(result_json.content)
            
            # Verifying the id for the object
            for item in result_json['value']:
                if item['name'] == folder_match:
                    return item['id']
            return None
        
        except Exception as e:
            print(f'Error: {e}')
    

    def get_drive_id(self, main_sharepoint_directory = 'Documents', folder_match = None):
        """
        : param: main_sharepoint_directory (str): represents the primary location for fetching the information.
            Usually, this information is present in a location called 'Documents', the default value.
        
        : param: folder_match (str): path within the main directory. It is the folder that will indicate 
            where the file must be located, within any levels that may have. Example: if you want to access a folder 
            called "SPC data" in the 'Documents' directory, folder_match = 'SPC data'. 
            If you want to access the content directly from 'Documents' root, keep folder_match = None
        """
        import requests
        from dotenv import load_dotenv

        load_dotenv()
        
        try: 
            headers = self.get_token()
            response = requests.get(self.url, headers = headers)
            drive_id = self.get_response_id(result_json = response, folder_match = main_sharepoint_directory)

            
            if folder_match: # run if it is not None        
                drive_response = requests.get(self.url + f'{drive_id}/root/children', headers = headers)
                root_folder_id = self.get_response_id(result_json = drive_response, folder_match = folder_match)
                
            return drive_id, root_folder_id
        
        except Exception as e:
            print(f'Error: {e}')


    def find_file(self, drive_id: str, folder_id: str, target_file_name: str, headers):
        """
            find_file is an API scraping function to find the ID of the target file on sharepoint.
            
            The arguments of this function are:
            - drive_id: string, the ID of the root folder on the API.
            - folder_id: string, uses the IDs of the folders from the root folder to search the target file.
            - target_file_name: string, the name of file to be searched. If found, stops the search.
            - headers: JSON, the authorization to run the get() request method.
            
            If the name is not match as required, the return is None.
        """
        import requests
        from dotenv import load_dotenv

        load_dotenv()

        try:
            # Requests to search the file ID.
            folder_url = f"{self.url}/{drive_id}/items/{folder_id}/children"
            folder_response = requests.get(folder_url, headers = headers)
            folder_result = folder_response.json()
            
            # This loops looks for the Download URL. If do not finds the file in the loop, it looks inside each folder.
            for item in folder_result['value']:
                if '@microsoft.graph.downloadUrl' in item:  # Check if it is a valid file
                    if item['name'] == target_file_name:
                        return item['id']
                else:
                    if 'folder' in item:
                        sub_folder_id = item['id']
                        file_id = self.find_file(drive_id, sub_folder_id, target_file_name, headers)
                        if file_id:
                            return file_id
            return None
        
        except Exception as e:
            print(f'Error: {e}')


    def download_file(self, target_file_name, main_sharepoint_directory = 'Documents', folder_match = None):
        """
            : param: main_sharepoint_directory (str): represents the primary location for fetching the information.
            Usually, this information is present in a location called 'Documents', the default value.
        
            : param: folder_match (str): path within the main directory. It is the folder that will indicate 
                where the file must be located, within any levels that may have. Example: if you want to access a folder 
                called "SPC data" in the 'Documents' directory, folder_match = 'SPC data'. 
                If you want to access the content directly from 'Documents' root, keep folder_match = None

            The function is able to read and download the sharepoint hosted file from the result given the file ID obtained by the method find_file()
            
            The arguments of this function are:
            - target_file_name: string, the name of file to be searched. If found, stops the search.
            - match: match the root folder of the application, after the documents folder is already accessed.
            
            If the name is not match as required, the return is None.
        """
        import requests
        from urllib.request import urlretrieve
        from dotenv import load_dotenv

        load_dotenv()

        try:
            # Getting each important ID to download the file
            headers = self.get_token()
            drive_id, root_folder_id = self.get_drive_id(main_sharepoint_directory, folder_match)
            file_id = self.find_file(drive_id, root_folder_id, target_file_name, headers)
            
            # Requests the Download URL from the API and downloads it within the system set up.
            if file_id:
                file_url = f"{self.url}/{drive_id}/items/{file_id}"
                file_result = requests.get(file_url, headers = headers).json()
                file_download_url = file_result["@microsoft.graph.downloadUrl"]
                urlretrieve(file_download_url, file_result['name'])
                print("Download successful!")
            else:
                raise InvalidInputsError("File not found.")
                
        except Exception as e:
            print(f'Error: {e}')
        # sharepoint_downloader = SharePointDownloader()
        # sharepoint_downloader.download_file(target_file_name = "data.xlsx")

